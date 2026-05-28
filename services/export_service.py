from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional, Sequence
from typing import Any, Mapping, Sequence

from config import EXPORT_DIR
from services.check_service import CheckService
from services.expense_service import ExpenseService
from services.registration_service import RegistrationService
from utils.date_utils import today_jalali


logger = logging.getLogger(__name__)


class ExportService:
    def __init__(
        self,
        check_service: CheckService,
        expense_service: Optional[ExpenseService] = None,
        registration_service: Optional[RegistrationService] = None,
    ):
        self.check_service = check_service
        self.expense_service = expense_service
        self.registration_service = registration_service

    def export_checks_to_excel(
        self,
        file_path: Optional[Path] = None,
        status: Optional[str] = None,
        from_due_date: Optional[str] = None,
        to_due_date: Optional[str] = None,
    ) -> Path:
        excel = self._require_excel_support()
        wb = excel['Workbook']()
        ws = wb.active
        ws.title = 'چک‌ها'

        headers = [
            'ID',
            'شماره صیادی چک 16 رقم',
            'شماره سریال چک',
            'نام ثبتنام کننده',
            'بانک',
            'صاحب حساب',
            'دریافت‌کننده',
            'مبلغ (ریال)',
            'تاریخ صدور',
            'تاریخ سررسید',
            'وضعیت',
            'توضیحات',
        ]

        styles = self._build_styles(excel)
        widths = self._append_header(ws, headers, styles)
        row_count = 0

        for check in self.check_service.iter_checks(
            status=status,
            from_due_date=from_due_date,
            to_due_date=to_due_date,
            batch_size=1000,
        ):
            values = [
                check.id,
                check.serial_18,
                check.serial_7,
                check.registrant_name,
                check.bank_name,
                check.account_owner,
                check.payee_name,
                int(check.amount or 0),
                check.issue_date,
                check.due_date,
                check.status,
                check.notes,
            ]
            self._append_data_row(ws, values, widths, styles)
            row_count += 1

        self._finalize_sheet(
            worksheet=ws,
            widths=widths,
            styles=styles,
            excel=excel,
            center_columns={1, 2, 3, 8, 9, 10, 11},
            amount_columns={8},
        )

        destination = self._resolve_file_path(file_path, prefix='checks_export')
        wb.save(destination)
        logger.info('Checks exported to Excel: rows=%s path=%s', row_count, destination)
        return destination

    def export_expenses_to_excel(
        self,
        file_path: Optional[Path] = None,
        search_text: str = '',
        category_id: Optional[int] = None,
        payment_method: str = '',
        from_date: str = '',
        to_date: str = '',
        linked_only: bool = False,
    ) -> Path:
        if self.expense_service is None:
            raise ValueError('Expense service is not configured.')

        excel = self._require_excel_support()
        styles = self._build_styles(excel)
        wb = excel['Workbook']()

        details_sheet = wb.active
        details_sheet.title = 'هزینه ها'

        detail_headers = [
            'ID',
            'عنوان هزینه',
            'دسته بندی',
            'روش پرداخت',
            'مبلغ (ریال)',
            'تاریخ هزینه',
            'مرجع چک',
            'فروشنده / پرداخت شونده',
            'توضیحات',
        ]
        detail_widths = self._append_header(details_sheet, detail_headers, styles)

        today_text = today_jalali().strftime('%Y/%m/%d')
        month_prefix = today_text[:7]

        rows_count = 0
        total_amount = 0
        today_amount = 0
        month_amount = 0
        active_days: set[str] = set()
        category_totals: dict[str, int] = {}
        method_totals: dict[str, int] = {}

        for expense in self.expense_service.iter_expenses(
            search_text=search_text,
            category_id=category_id,
            payment_method=payment_method,
            from_date=from_date,
            to_date=to_date,
            batch_size=1000,
        ):
            if linked_only and expense.reference_check_id is None:
                continue

            amount = int(expense.amount or 0)
            method_label = self.expense_service.PAYMENT_METHODS.get(
                expense.payment_method,
                expense.payment_method,
            )

            values = [
                expense.id,
                expense.title,
                expense.category_name,
                method_label,
                amount,
                expense.expense_date,
                expense.reference_check_id or '',
                expense.vendor,
                expense.notes,
            ]
            self._append_data_row(details_sheet, values, detail_widths, styles)

            rows_count += 1
            total_amount += amount
            if expense.expense_date:
                active_days.add(expense.expense_date)
            if expense.expense_date == today_text:
                today_amount += amount
            if (expense.expense_date or '').startswith(month_prefix):
                month_amount += amount

            category_name = expense.category_name or 'نامشخص'
            category_totals[category_name] = category_totals.get(category_name, 0) + amount
            method_totals[method_label] = method_totals.get(method_label, 0) + amount

        average_daily = int(total_amount / len(active_days)) if active_days else 0
        summary_rows = [
            ('تعداد رکورد', rows_count),
            ('مجموع هزینه', total_amount),
            ('هزینه امروز', today_amount),
            ('هزینه ماه جاری', month_amount),
            ('میانگین روزانه', average_daily),
        ]

        top_categories = sorted(category_totals.items(), key=lambda item: item[1], reverse=True)[:5]
        methods = sorted(method_totals.items(), key=lambda item: item[1], reverse=True)

        summary_sheet = wb.create_sheet('خلاصه گزارش')
        summary_widths = self._append_header(summary_sheet, ['شاخص', 'مقدار'], styles)
        for metric, value in summary_rows:
            self._append_data_row(summary_sheet, [metric, value], summary_widths, styles)
        self._finalize_sheet(summary_sheet, summary_widths, styles, excel, {2}, {2})

        category_sheet = wb.create_sheet('هزینه به تفکیک دسته')
        category_widths = self._append_header(category_sheet, ['دسته بندی', 'مجموع هزینه (ریال)'], styles)
        for name, amount in top_categories:
            self._append_data_row(category_sheet, [name, amount], category_widths, styles)
        self._finalize_sheet(category_sheet, category_widths, styles, excel, {2}, {2})

        methods_sheet = wb.create_sheet('هزینه به تفکیک پرداخت')
        methods_widths = self._append_header(methods_sheet, ['روش پرداخت', 'مجموع هزینه (ریال)'], styles)
        for name, amount in methods:
            self._append_data_row(methods_sheet, [name, amount], methods_widths, styles)
        self._finalize_sheet(methods_sheet, methods_widths, styles, excel, {2}, {2})

        self._finalize_sheet(
            details_sheet,
            detail_widths,
            styles,
            excel,
            center_columns={1, 4, 5, 6, 7},
            amount_columns={5},
        )

        destination = self._resolve_file_path(file_path, prefix='expenses_export')
        wb.save(destination)
        logger.info('Expenses exported to Excel: rows=%s path=%s', rows_count, destination)
        return destination

    def export_registrations_to_excel(
        self,
        file_path: Optional[Path] = None,
        search_text: str = '',
    ) -> Path:
        if self.registration_service is None:
            raise ValueError('Registration service is not configured.')

        excel = self._require_excel_support()
        styles = self._build_styles(excel)
        wb = excel['Workbook']()

        registrations = self.registration_service.list_all(search_text=search_text)

        registrations_sheet = wb.active
        registrations_sheet.title = 'ثبت نام ها'
        registration_headers = [
            'شناسه ثبت نام',
            'نام مشتری',
            'کد ملی',
            'شماره تماس',
            'دوره',
            'تاریخ ثبت نام',
            'روش پرداخت',
            'شهریه کل (ریال)',
            'پرداخت اولیه (ریال)',
            'درآمد ثبت شده (ریال)',
            'مانده (ریال)',
            'مبلغ چکی (ریال)',
            'مبلغ غیرچکی (ریال)',
            'تعداد پرداخت',
            'تعداد چک لینک شده',
            'توضیحات',
        ]
        registration_widths = self._append_header(registrations_sheet, registration_headers, styles)

        method_labels = self.registration_service.PAYMENT_METHOD_LABELS
        for registration in registrations:
            checks_total = int(getattr(registration, 'check_income_total', 0) or 0)
            initial_payment = int(getattr(registration, 'initial_payment', 0) or 0)
            remaining_balance = int(registration.total_fee or 0) - initial_payment - checks_total
            registrations_sheet.append(
                [
                    registration.id,
                    registration.customer_name,
                    registration.national_code,
                    registration.phone,
                    registration.course_name,
                    registration.registration_date,
                    method_labels.get(registration.payment_method, registration.payment_method),
                    int(registration.total_fee or 0),
                    initial_payment,
                    int(registration.income_total or 0),
                    remaining_balance,
                    checks_total,
                    int(getattr(registration, 'non_check_income_total', 0) or 0),
                    int(registration.payment_count or 0),
                    int(registration.check_count or 0),
                    registration.description,
                ]
            )
            self._apply_row_borders(registrations_sheet, styles, registration_widths)

        self._finalize_sheet(
            worksheet=registrations_sheet,
            widths=registration_widths,
            styles=styles,
            excel=excel,
            center_columns={1, 6, 7, 14, 15},
            amount_columns={8, 9, 10, 11, 12, 13},
        )

        payments_sheet = wb.create_sheet('پرداخت های لینک شده')
        payment_headers = [
            'شناسه پرداخت',
            'شناسه ثبت نام',
            'نام مشتری',
            'روش ثبت پرداخت',
            'مبلغ (ریال)',
            'تاریخ پرداخت',
            'شناسه چک',
            'شماره سریال چک',
            'سررسید چک',
            'توضیحات',
        ]
        payment_widths = self._append_header(payments_sheet, payment_headers, styles)

        payments_count = 0
        for registration in registrations:
            for payment in self.registration_service.list_payments(int(registration.id or 0)):
                payment_method = str(payment['payment_method'] or '').upper()
                if payment_method == 'CHECK':
                    rendered_method = 'چک'
                else:
                    non_check_method = str(
                        payment['registration_payment_method'] or registration.payment_method
                    ).upper()
                    rendered_method = method_labels.get(
                        non_check_method,
                        non_check_method,
                    )
                    rendered_method = f'غیرچکی ({rendered_method})'
                payments_sheet.append(
                    [
                        int(payment['id'] or 0),
                        int(payment['registration_id'] or 0),
                        registration.customer_name,
                        rendered_method,
                        int(payment['amount'] or 0),
                        payment['payment_date'] or '',
                        payment['check_id'] or '',
                        payment['serial_7'] or '',
                        payment['due_date'] or '',
                        payment['notes'] or '',
                    ]
                )
                self._apply_row_borders(payments_sheet, styles, payment_widths)
                payments_count += 1

        self._finalize_sheet(
            worksheet=payments_sheet,
            widths=payment_widths,
            styles=styles,
            excel=excel,
            center_columns={1, 2, 4, 6, 7, 8, 9},
            amount_columns={5},
        )

        financials = self.registration_service.build_financial_report()
        summary_sheet = wb.create_sheet('خلاصه مالی')
        summary_widths = self._append_header(summary_sheet, ['شاخص', 'مقدار'], styles)
        summary_rows = [
            ('تعداد ثبت نام های گزارش', len(registrations)),
            ('تعداد پرداخت های لینک شده', payments_count),
            ('مجموع درآمد', int(financials.get('total_income', 0))),
            ('مجموع هزینه ها', int(financials.get('total_expenses', 0))),
            ('خالص درآمد', int(financials.get('net_income', 0))),
        ]
        for title, amount in summary_rows:
            summary_sheet.append([title, amount])
            self._apply_row_borders(summary_sheet, styles, summary_widths)

        self._finalize_sheet(
            worksheet=summary_sheet,
            widths=summary_widths,
            styles=styles,
            excel=excel,
            center_columns={2},
            amount_columns={2},
        )

        destination = self._resolve_file_path(file_path, prefix='registrations_export')
        wb.save(destination)
        logger.info(
            'Registrations exported to Excel: registrations=%s payments=%s path=%s',
            len(registrations),
            payments_count,
            destination,
        )
        return destination

    def export_dashboard_analytics_to_excel(
        self,
        file_path: Optional[Path] = None,
        payload: Optional[Mapping[str, object]] = None,
    ) -> Path:
        data = dict(payload or {})
        excel = self._require_excel_support()
        styles = self._build_styles(excel)
        wb = excel['Workbook']()

        summary_sheet = wb.active
        summary_sheet.title = 'خلاصه داشبورد'
        summary_widths = self._append_header(summary_sheet, ['شاخص', 'مقدار'], styles)

        summary = data.get('summary') or {}
        growth = data.get('growth') or {}
        summary_rows = [
            ('بازه گزارش', data.get('period_label') or '-'),
            ('کل درآمد', int(summary.get('incomes', 0) or 0)),
            ('کل هزینه', int(summary.get('expenses', 0) or 0)),
            ('خالص', int(summary.get('net', 0) or 0)),
            ('تعداد ثبت نام', int(summary.get('registrations_count', 0) or 0)),
            ('تعداد چک', int(summary.get('checks_count', 0) or 0)),
            ('مبلغ چک ها', int(summary.get('checks_amount', 0) or 0)),
            ('رشد درآمد', growth.get('incomes') or '-'),
            ('رشد هزینه', growth.get('expenses') or '-'),
            ('رشد ثبت نام', growth.get('registrations') or '-'),
            ('رشد چک', growth.get('checks') or '-'),
        ]
        for key, value in summary_rows:
            self._append_data_row(summary_sheet, [key, value], summary_widths, styles)
        self._finalize_sheet(summary_sheet, summary_widths, styles, excel, {2}, {2})

        monthly_sheet = wb.create_sheet('تحلیل ماهانه')
        monthly_headers = ['دوره', 'درآمد', 'هزینه', 'خالص', 'تعداد ثبت نام', 'تعداد چک', 'مبلغ چک']
        monthly_widths = self._append_header(monthly_sheet, monthly_headers, styles)
        for row in data.get('monthly_analytics') or []:
            self._append_data_row(
                monthly_sheet,
                [
                    row.get('period', ''),
                    int(row.get('income', 0) or 0),
                    int(row.get('expense', 0) or 0),
                    int(row.get('net', 0) or 0),
                    int(row.get('registrations', 0) or 0),
                    int(row.get('checks_count', 0) or 0),
                    int(row.get('checks_amount', 0) or 0),
                ],
                monthly_widths,
                styles,
            )
        self._finalize_sheet(monthly_sheet, monthly_widths, styles, excel, {1, 5, 6}, {2, 3, 4, 7})

        self._append_generic_table_sheet(
            wb,
            excel,
            styles,
            'مقایسه دوره',
            data.get('monthly_comparison') or [],
        )
        self._append_generic_table_sheet(
            wb,
            excel,
            styles,
            'تفکیک هزینه',
            data.get('category_breakdown') or [],
        )
        self._append_generic_table_sheet(
            wb,
            excel,
            styles,
            'هشدارهای ریسک چک',
            data.get('upcoming_risks') or [],
        )
        self._append_ai_report_sheet(
            wb,
            excel,
            styles,
            data.get('ai_report') or {},
        )

        destination = self._resolve_file_path(file_path, prefix='dashboard_analytics')
        wb.save(destination)
        logger.info('Dashboard analytics exported to Excel: path=%s', destination)
        return destination

    def _append_generic_table_sheet(
        self,
        workbook,
        excel: Mapping[str, object],
        styles: Mapping[str, object],
        sheet_name: str,
        rows: Sequence[Mapping[str, object]],
    ):
        sheet = workbook.create_sheet(sheet_name)
        if not rows:
            widths = self._append_header(sheet, ['اطلاعات'], styles)
            self._append_data_row(sheet, ['داده ای برای نمایش وجود ندارد.'], widths, styles)
            self._finalize_sheet(sheet, widths, styles, excel, set(), set())
            return

        headers = list(rows[0].keys())
        widths = self._append_header(sheet, headers, styles)
        for row in rows:
            self._append_data_row(sheet, [row.get(header, '') for header in headers], widths, styles)
        self._finalize_sheet(sheet, widths, styles, excel, set(), set())

    def _append_ai_report_sheet(
        self,
        workbook,
        excel: Mapping[str, object],
        styles: Mapping[str, object],
        ai_report: Mapping[str, object],
    ):
        if not ai_report:
            return

        sheet = workbook.create_sheet('AI Structured Report')
        widths = self._append_header(sheet, ['Section', 'Content'], styles)
        self._append_data_row(sheet, ['Summary', ai_report.get('summary', '-')], widths, styles)

        key_metrics = ai_report.get('key_metrics') or {}
        for key, value in key_metrics.items():
            self._append_data_row(sheet, [f'Key Metric - {key}', value], widths, styles)

        for risk in ai_report.get('risks') or []:
            self._append_data_row(
                sheet,
                [
                    'Risk',
                    f"{risk.get('title', '-')}"
                    f" | Severity={risk.get('severity', '-')}"
                    f" | Count={risk.get('count', 0)}",
                ],
                widths,
                styles,
            )

        for trend in ai_report.get('trends') or []:
            self._append_data_row(
                sheet,
                ['Trend', f"{trend.get('metric', '-')}={trend.get('change_pct', 0)}%"],
                widths,
                styles,
            )

        for item in ai_report.get('recommendations') or []:
            self._append_data_row(sheet, ['Recommendation', item], widths, styles)

        for item in ai_report.get('reasoning_steps') or []:
            self._append_data_row(sheet, ['Reasoning Step', item], widths, styles)

        self._finalize_sheet(sheet, widths, styles, excel, set(), set())

    def _resolve_file_path(self, file_path: Optional[Path], prefix: str) -> Path:
        if file_path is None:
            EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            return EXPORT_DIR / f'{prefix}_{timestamp}.xlsx'
        return Path(file_path)

    @staticmethod
    def _require_excel_support() -> dict[str, object]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                'Excel export requires openpyxl. Install dependencies from requirements.txt first.'
            ) from exc

        return {
            'Workbook': Workbook,
            'Alignment': Alignment,
            'Border': Border,
            'Font': Font,
            'PatternFill': PatternFill,
            'Side': Side,
            'get_column_letter': get_column_letter,
        }

    @staticmethod
    def _build_styles(excel: dict[str, object]) -> dict[str, object]:
        PatternFill = excel['PatternFill']
        Font = excel['Font']
        Side = excel['Side']
        Border = excel['Border']
        Alignment = excel['Alignment']

        return {
            'header_fill': PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78'),
            'header_font': Font(color='FFFFFF', bold=True),
            'thin_border': Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9'),
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center', wrap_text=True),
        }

    def _append_header(self, worksheet, headers: Sequence[str], styles: dict[str, object]) -> dict[int, int]:
        worksheet.append(list(headers))
        widths: dict[int, int] = {}
        for index, value in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=index)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['thin_border']
            cell.alignment = styles['center']
            widths[index] = max(10, min(60, len(str(value)) + 4))
        return widths

    def _append_data_row(
        self,
        worksheet,
        values: Iterable[object],
        widths: dict[int, int],
        styles: dict[str, object],
    ):
        worksheet.append(list(values))
        row_index = worksheet.max_row
        for col_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.border = styles['thin_border']
            rendered = '' if value is None else str(value)
            widths[col_index] = max(widths.get(col_index, 10), min(70, len(rendered) + 2))

    @staticmethod
    def _apply_row_borders(worksheet, styles: dict[str, object], widths: dict[int, int]):
        row_index = worksheet.max_row
        max_col = worksheet.max_column
        for col_index in range(1, max_col + 1):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.border = styles['thin_border']
            rendered = '' if cell.value is None else str(cell.value)
            widths[col_index] = max(widths.get(col_index, 10), min(70, len(rendered) + 2))

    def _finalize_sheet(
        self,
        worksheet,
        widths: dict[int, int],
        styles: dict[str, object],
        excel: dict[str, object],
        center_columns: set[int],
        amount_columns: set[int],
    ):
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column in center_columns:
                    cell.alignment = styles['center']
                else:
                    cell.alignment = styles['left']

                if cell.column in amount_columns and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

        get_column_letter = excel['get_column_letter']
        for col_index, width in widths.items():
            worksheet.column_dimensions[get_column_letter(col_index)].width = float(max(10, min(width, 70)))

        worksheet.sheet_view.rightToLeft = True
